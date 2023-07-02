import numpy as np
import math
from tqdm import tqdm

import base64
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ocr.utils.text_utils import ocr, get_text_bounding_box
from ocr.utils.text_region_detection import *
from ocr.utils.text_line_detection import text_line_detection
from ocr.utils.load_models import YOLOv5_model, CRAFT_model, VietOCR_model
from ocr.utils.table_structure_postprocess import objects_to_cells
import ocr.utils.table_config as table_config
from ocr.preprocess import deskew, remove_line
from ocr.path import TABLE_STRUCTURE_RECOGNITION_PATH

def get_table_text_object(img, net):
    boxes = get_text_bounding_box(img, net)
    tokens_in_table = [np.array([box[0].astype('int32'), box[2].astype('int32')]).flatten().tolist() for box in boxes]
    return tokens_in_table

def table_structure_recognition(image, model):
    print('TABLE STRUCTURE DETECTION')
    pred = model(image, size=640)
    pred = pred.xywhn[0]    
    result = pred.numpy()
    return result

def convert_structure(img, tsr_result, tokens_in_table):
    print('CLEAN TABLE STRUCTURE')
    width = img.shape[1]
    height = img.shape[0]
    bboxes = []
    scores = []
    labels = []
    for item in tsr_result:
        class_id = int(item[5])
        score = float(item[4])
        min_x = item[0]
        min_y = item[1]
        w = item[2]
        h = item[3]
        
        x1 = int((min_x-w/2)*width)
        y1 = int((min_y-h/2)*height)
        x2 = int((min_x+w/2)*width)
        y2 = int((min_y+h/2)*height)

        bboxes.append([x1, y1, x2, y2])
        scores.append(score)
        labels.append(class_id)

    table_objects = []
    for bbox, score, label in zip(bboxes, scores, labels):
        table_objects.append({'bbox': bbox, 'score': score, 'label': label})

    _, cells, _, table_rows_cols = objects_to_cells(table_objects, tokens_in_table, table_config.structure_class_names, table_config.structure_class_thresholds)
    return cells, table_rows_cols

def ocr_cell(img, min_line_height, detector):
    result = ''
    # text_cell_metadata = []
    text_metadata = text_region_detection(img, dilate_kernel=(7,7))
    for metadata in text_metadata:
        x_text = metadata['text-region'][0]
        y_text = metadata['text-region'][1]
        w_text = metadata['text-region'][2]
        h_text = metadata['text-region'][3]
        text_region = img[y_text:y_text+h_text, x_text:x_text+w_text]
        lines = text_line_detection(text_region, min_line_height)
        # lines_metadata = []
        for line in lines:
            x_line = line[0]
            y_line = line[1]
            w_line = line[2]
            h_line = line[3]
            line_region = text_region[y_line:y_line+h_line, x_line:x_line+w_line]
            new_x_line, new_w_line = preprocess_line_region(line_region)
            new_line_region = text_region[y_line:y_line+h_line, new_x_line:new_x_line+new_w_line]
            text = ocr(new_line_region, detector)
            # lines_metadata.append({'line_coordinates': [int(new_x_line), int(y_line), int(new_w_line), int(h_line)], 'text': text})
            result+=text+' '
        # text_cell_metadata.append({'region_coordinates': [int(x_text), int(y_text), int(w_text), int(h_text)], 'lines': lines_metadata, 'text': result})

    # return text_cell_metadata
    return result

def ocr_table_cells(img, cells, table_rows_cols, ocr_model):
    print('TABLE OCR')
    min_text_height = 999
    for cell in tqdm(cells):
        text = ''
        if len(cell['spans'])!=0:
            min_text_height = min([bb[3]-bb[1] for bb in cell['spans']]+[min_text_height])
            # Calculate min line height by using text height
            min_line_height = min([bb[3]-bb[1] for bb in cell['spans']])/4
            bbox = cell['bbox']
            x1 = int(bbox[0])
            y1 = int(bbox[1])
            x2 = int(bbox[2])
            y2 = int(bbox[3])
            crop = img[y1:y2, x1:x2]
            if not (crop.shape[0]==0 or crop.shape[1]==0):
                text = ocr_cell(crop, min_line_height, ocr_model)
        cell['text'] = text

    row_height_data = []
    column_width_data = []

    excel_font_size = 18 # in pixels, in point (pt) is 14
    ratio = excel_font_size/min_text_height
    for row in table_rows_cols['rows']:
        row_height_data.append(math.ceil((row['bbox'][3]-row['bbox'][1]) * ratio))
    for col in table_rows_cols['cols']:
        column_width_data.append(math.ceil((col['bbox'][2]-col['bbox'][0]) * ratio))
    return cells, row_height_data, column_width_data

def create_excel_base64(cells, row_heights, column_widths, id):
    # Tạo workbook mới
    workbook = Workbook()
    sheet = workbook.active
    font = Font(name='Arial', size = 14)
    # Thiết lập chiều cao hàng và chiều dài cột (Từ pixel chuyển về đơn vị tính của Excel)
    for row, height in enumerate(row_heights, start=1):
        sheet.row_dimensions[row].height = height*3/4

    for col, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width*1/7

    # Đổ dữ liệu vào từng cell
    for cell_data in cells:
        row_nums = cell_data['row_nums']
        col_nums = cell_data['column_nums']
        text = cell_data['text']

        # Merge cell nếu cần thiết
        if len(row_nums) > 1 or len(col_nums) > 1:
            start_row = row_nums[0] + 1
            end_row = row_nums[-1] + 1
            start_col = col_nums[0] + 1
            end_col = col_nums[-1] + 1
            sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=start_col, end_column=end_col)

            # Đặt giá trị và căn chỉnh cho cell merge
            merge_cell = sheet.cell(row=start_row, column=start_col)
            merge_cell.value = text
            merge_cell.alignment = Alignment(wrapText=True)
            merge_cell.font = font
        else:
            # Đặt giá trị và căn chỉnh cho cell không merge
            row = row_nums[0] + 1
            col = col_nums[0] + 1
            cell = sheet.cell(row=row, column=col)
            cell.value = text
            cell.alignment = Alignment(wrapText=True)
            cell.font = font

    # Tạo BytesIO để lưu workbook vào bộ nhớ
    # excel_data = BytesIO()
    # workbook.save(excel_data)
    workbook.save('temp/{}.xlsx'.format(id))
    workbook = load_workbook(filename='temp/{}.xlsx'.format(id))
    excel_data = BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)

    # Chuyển đổi dữ liệu Excel thành base64
    base64_data = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,'+base64.b64encode(excel_data.read()).decode('utf-8')

    return base64_data

# def convert_cells_to_luckysheet(cells, table_rows_cols, min_text_height):
#     cell_data = []
#     merge_data = {}
#     row_height_data = {}
#     column_width_data = {}

#     luckysheet_font_size = 14
#     ratio = luckysheet_font_size/min_text_height 
#     for idx, row in enumerate(table_rows_cols['rows']):
#         row_height_data[idx] = math.ceil((row['bbox'][3]-row['bbox'][1]) * ratio)
#     for idx, col in enumerate(table_rows_cols['cols']):
#         column_width_data[idx] = math.ceil((col['bbox'][2]-col['bbox'][0]) * ratio)

#     # Process cells data
#     for cell in cells:
#         min_row = min(cell['row_nums'])
#         max_row = max(cell['row_nums'])
#         min_column = min(cell['column_nums'])
#         max_column = max(cell['column_nums'])

#         # Update handsontable data with cell text
#         for row in range(min_row, max_row + 1):
#             for column in range(min_column, max_column + 1):
#                 # Single cell
#                 if min_row == max_row and min_column == max_column:
#                     cell_data.append({
#                         'r': row,
#                         'c': column,
#                         'v': {
#                             'v': cell['text'],
#                             'm': cell['text'],
#                             'ct': { 'fa': "@", 't': "s" },
#                             'bg': None,
#                             'bl': 0,
#                             'it': 0,
#                             'ff': 0,
#                             'fs': luckysheet_font_size,
#                             'fc': 'rgb(0, 0, 0)',
#                             'ht': 1,
#                             'vt': 1,
#                             'tb': "2"
#                         }
#                     })
#                 # Merge cell
#                 else:
#                     if (row==min_row and column==min_column):
#                         cell_data.append({
#                             'r': row,
#                             'c': column,
#                             'v': {
#                                 'mc': {
#                                     'r': min_row,
#                                     'c': min_column,
#                                     'rs': max_row - min_row + 1,
#                                     'cs': max_column - min_column + 1
#                                 },
#                                 'v': cell['text'],
#                                 'm': cell['text'],
#                                 'ct': { 'fa': "@", 't': "s" },
#                                 'bg': None,
#                                 'bl': 0,
#                                 'it': 0,
#                                 'ff': 0,
#                                 'fs': luckysheet_font_size,
#                                 'fc': 'rgb(0, 0, 0)',
#                                 'ht': 1,
#                                 'vt': 1,
#                                 'tb': "2"
#                             }
#                         })
#                     else:
#                         cell_data.append({
#                             'r': row,
#                             'c': column,
#                             'v': {
#                                 'mc': {
#                                     'r': min_row,
#                                     'c': min_column,
#                                     'rs': max_row - min_row + 1,
#                                     'cs': max_column - min_column + 1
#                                 }
#                             }
#                         })


#         # Check if cell is a merged cell
#         if min_row != max_row or min_column != max_column:
#             merge_info = {
#                 'r': min_row,
#                 'c': min_column,
#                 'rs': max_row - min_row + 1,
#                 'cs': max_column - min_column + 1
#             }
#             merge_data['{}_{}'.format(min_row, min_column)] = merge_info

#     luckysheet = {'celldata': cell_data, 'merge': merge_data, "rowlen": row_height_data, 'columnlen': column_width_data}

#     return luckysheet

def ocr_table(img, tsr_model, net, detector, id):
    print('LOAD MODELS')

    img = deskew(img)
    tokens_in_table = get_table_text_object(img, net)
    tsr_result = table_structure_recognition(img, tsr_model)
    cells, table_rows_cols= convert_structure(img, tsr_result, tokens_in_table)
    img = remove_line(img)
    text_cells, row_heights, column_widths = ocr_table_cells(img, cells, table_rows_cols, detector)
    # metadata = convert_cells_to_luckysheet(text_cells, table_rows_cols, min_text_height)
    base64_data = create_excel_base64(text_cells, row_heights, column_widths, id)

    return base64_data


